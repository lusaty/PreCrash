import os
import json
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

def read_jsonl(file_path):
    """
    读取JSONL文件并返回一个列表，其中每个元素是一个JSON对象。
    """
    data = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line_number, line in enumerate(f, 1):
            try:
                data.append(json.loads(line))
            except json.JSONDecodeError as e:
                print(f"JSON解码错误在文件 {file_path} 的第 {line_number} 行: {line.strip()}")
    return data

def extract_report_id(report):
    """
    从报告中提取唯一标识符，用于匹配可重现和不可重现报告。
    假设每个报告中有一个唯一的ID字段，如 'id' 或 'report_id'。
    需要根据实际数据结构调整。
    """
    report_id = report.get('id') or report.get('report_id') or report.get('issue_number')
    if not report_id:
        print(f"报告缺少唯一 ID 字段: {report}")
    return report_id

def extract_title(report):
    """
    从报告中提取标题。
    根据实际数据结构调整字段名。
    """
    return report.get('title', 'Unknown')

def extract_creation_time(report):
    """
    从报告中提取创建时间。
    根据实际数据结构调整字段名。
    """
    return report.get('created_at', 'Unknown')

def process_project(project_path):
    """
    处理单个项目，读取崩溃报告并打标签，提取所需属性，并去重同一项目下重复的issue。
    """
    crash_reports_path = os.path.join(project_path, 'crash_reports')
    reproducible_reports_path = os.path.join(project_path, 'reproducible_crash_reports')

    # 检查 crash_reports 文件夹是否存在
    if not os.path.exists(crash_reports_path):
        print(f"警告：项目 '{project_path}' 中缺少 'crash_reports' 文件夹。跳过该项目。")
        return []

    # 获取所有 crash_reports 文件
    crash_report_files = [f for f in os.listdir(crash_reports_path) if f.endswith('_crash_reports.jsonl')]

    all_reports = {}
    reproducible_ids = set()

    # 提取仓库名称（假设项目文件夹名即为仓库名称）
    repository_name = os.path.basename(project_path)

    # 读取所有可重现的报告ID
    reproducible_files = [f for f in os.listdir(reproducible_reports_path) if f.endswith('_reproducible_crash_reports.jsonl')]
    for file in reproducible_files:
        file_path = os.path.join(reproducible_reports_path, file)
        reproducible_data = read_jsonl(file_path)
        for report in reproducible_data:
            report_id = extract_report_id(report)
            if report_id:
                reproducible_ids.add(report_id)

    # 读取所有崩溃报告并打标签
    for file in crash_report_files:
        file_path = os.path.join(crash_reports_path, file)
        crash_data = read_jsonl(file_path)
        for report in crash_data:
            report_id = extract_report_id(report)
            if not report_id:
                print(f"跳过没有ID的报告：文件 {file_path} 中的报告")
                continue  # 跳过没有ID的报告
            # 生成全局唯一的 unique_id
            global_unique_id = f"{repository_name}_{report_id}"
            if global_unique_id in all_reports:
                print(f"跳过重复的报告ID {global_unique_id}：文件 {file_path}")
                continue  # 跳过重复的报告
            label = 1 if report_id in reproducible_ids else 0
            report['label'] = label
            # 提取其他属性
            report['title'] = extract_title(report)
            report['created_at'] = extract_creation_time(report)
            # 添加仓库名称
            report['repository'] = repository_name
            # 更新 'unique_id' 字段
            report['unique_id'] = global_unique_id
            # 添加到 all_reports 字典，确保唯一性
            all_reports[global_unique_id] = report

    print(f"项目 '{repository_name}' 处理完毕，共 {len(all_reports)} 个 reports，其中 {sum(r['label'] for r in all_reports.values())} 个可重现。")
    return list(all_reports.values())

def clean_excel_string(s):
    """
    移除字符串中所有非法的Excel字符。
    """
    if isinstance(s, str):
        # 移除控制字符，保留制表符、换行符和回车符
        return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', s)
    return s

def main(dataset_path, output_excel='labeled_crash_reports.xlsx'):
    """
    处理整个数据集，并保存标注后的报告到Excel文件。
    """
    all_projects = [os.path.join(dataset_path, d) for d in os.listdir(dataset_path)
                    if os.path.isdir(os.path.join(dataset_path, d))]

    all_labeled_reports = []

    for project in tqdm(all_projects, desc='Processing Projects'):
        labeled_reports = process_project(project)
        all_labeled_reports.extend(labeled_reports)

    if not all_labeled_reports:
        print("没有找到任何崩溃报告。")
        return

    # 转换为DataFrame
    df = pd.DataFrame(all_labeled_reports)

    # 根据实际需要选择需要的列
    columns_to_include = ['repository', 'unique_id', 'id', 'report_id', 'issue_number', 'title', 'created_at', 'label']
    existing_columns = [col for col in columns_to_include if col in df.columns]

    # 确保 'repository' 放在第一列
    if 'repository' in df.columns:
        df = df[['repository'] + [col for col in df.columns if col != 'repository']]
    else:
        df['repository'] = 'Unknown'
        df = df[['repository'] + [col for col in df.columns if col != 'repository']]

    # 去除全局重复的 unique_ids（现在是全局唯一的）
    before_dedup = len(df)
    if 'unique_id' in df.columns:
        df = df.drop_duplicates(subset=['unique_id'], keep='first')
    else:
        print("警告：DataFrame 中缺少 'unique_id' 列，无法去重。")
    after_dedup = len(df)
    if before_dedup != after_dedup:
        print(f"去重前 issues 数量: {before_dedup}")
        print(f"去重后 issues 数量: {after_dedup}")

    # 清理 DataFrame 中的非法字符
    string_columns = df.select_dtypes(include=['object']).columns
    for col in string_columns:
        df[col] = df[col].apply(clean_excel_string)

    # 初始化Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Labeled Crash Reports"

    # 定义表头
    headers = df.columns.tolist()
    ws.append(headers)

    # 设置列宽（可选）
    column_widths = {
        "A": 20, "B": 30, "C": 15, "D": 15, "E": 50, "F": 25, "G": 10, "H": 10,
        # 根据实际列数和内容调整更多列宽
    }

    for i, column in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        if col_letter in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[col_letter]
        else:
            ws.column_dimensions[col_letter].width = 20  # 默认宽度

    # 写入数据
    for index, row in df.iterrows():
        ws.append(row.tolist())

    # 保存Excel文件
    wb.save(output_excel)
    print(f"标注完成，结果已保存到 {output_excel}")
    print(f"标注后的 issues 总数: {len(df)}")

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="处理并标注崩溃报告数据集。")
    parser.add_argument(
        "-d", "--dataset",
        type=str,
        required=True,
        help="数据集的根目录路径。"
    )
    parser.add_argument(
        "-o", "--output",
        type=str,
        default="datasets/crash_reports_with_labels.xlsx",
        help="输出的标注后Excel文件名。"
    )
    args = parser.parse_args()

    dataset_root = args.dataset
    output_file = args.output

    main(dataset_root, output_file)
